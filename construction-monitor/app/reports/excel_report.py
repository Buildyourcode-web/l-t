"""
Excel Report Generator using openpyxl.

Generates a two-sheet workbook:
- Sheet 1: Summary (statistics by camera and violation type)
- Sheet 2: Details (all violations with full data)
"""
from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from app.core.logging_config import get_logger

logger = get_logger(__name__, "api")

# Brand colors
HEADER_BG = "1E40AF"   # deep blue
HEADER_FG = "FFFFFF"   # white
ALT_ROW_BG = "F8FAFC"  # very light gray
FIRE_BG = "FEE2E2"     # light red for fire rows
HELMET_BG = "FEF3C7"   # light amber for helmet rows
VEST_BG = "D1FAE5"     # light green for vest rows


def _header_style() -> dict:
    return {
        "font": Font(bold=True, color=HEADER_FG, size=11),
        "fill": PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center"),
    }


def _apply_style(cell, font=None, fill=None, alignment=None) -> None:
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment


def _auto_width(ws) -> None:
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def generate_excel(
    violations: list,
    output_path: str,
    report_date: date,
) -> None:
    """
    Generate an Excel report for the given violations.

    Args:
        violations: List of Violation ORM objects
        output_path: Absolute path to save the .xlsx file
        report_date: The date this report covers
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Title
    ws_summary.merge_cells("A1:D1")
    title_cell = ws_summary["A1"]
    title_cell.value = f"Daily Safety Report — {report_date.strftime('%B %d, %Y')}"
    title_cell.font = Font(bold=True, size=14, color=HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center")
    ws_summary.row_dimensions[1].height = 30
    ws_summary.append([])

    # Stats by violation type
    ws_summary.append(["Violation Type", "Count", "% of Total"])
    for cell in ws_summary[ws_summary.max_row]:
        _apply_style(cell, **_header_style())

    type_counts = {}
    for v in violations:
        type_counts[v.violation_type] = type_counts.get(v.violation_type, 0) + 1

    total = len(violations)
    for vtype, cnt in sorted(type_counts.items(), key=lambda x: -x[1]):
        pct = f"{cnt / total * 100:.1f}%" if total > 0 else "0%"
        ws_summary.append([vtype.replace("_", " ").title(), cnt, pct])

    ws_summary.append([])

    # Stats by camera
    ws_summary.append(["Camera", "Violations", "Most Common Type"])
    for cell in ws_summary[ws_summary.max_row]:
        _apply_style(cell, **_header_style())

    cam_data = {}
    for v in violations:
        if v.camera_name not in cam_data:
            cam_data[v.camera_name] = {}
        cam_data[v.camera_name][v.violation_type] = cam_data[v.camera_name].get(v.violation_type, 0) + 1

    for cam, type_map in sorted(cam_data.items()):
        total_cam = sum(type_map.values())
        most_common = max(type_map, key=type_map.get)
        ws_summary.append([cam, total_cam, most_common.replace("_", " ").title()])

    _auto_width(ws_summary)
    ws_summary.freeze_panes = "A4"

    # ── Sheet 2: Details ──────────────────────────────────────────────────
    ws_detail = wb.create_sheet(title="Details")

    headers = ["#", "Date", "Time", "Camera", "Violation Type", "Track ID", "Confidence", "Screenshot Path"]
    ws_detail.append(headers)
    for cell in ws_detail[1]:
        _apply_style(cell, **_header_style())
    ws_detail.row_dimensions[1].height = 22
    ws_detail.freeze_panes = "A2"

    for idx, v in enumerate(violations, start=1):
        date_str = v.detected_at.strftime("%Y-%m-%d") if v.detected_at else ""
        time_str = v.detected_at.strftime("%H:%M:%S") if v.detected_at else ""
        conf_str = f"{v.confidence * 100:.1f}%"

        row = [
            idx,
            date_str,
            time_str,
            v.camera_name,
            v.violation_type.replace("_", " ").title(),
            v.track_id,
            conf_str,
            v.image_path or "",
        ]
        ws_detail.append(row)

        # Row color by violation type
        row_num = ws_detail.max_row
        if v.violation_type == "fire":
            bg = FIRE_BG
        elif v.violation_type == "no_helmet":
            bg = HELMET_BG
        else:
            bg = ALT_ROW_BG if idx % 2 == 0 else "FFFFFF"

        fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        for cell in ws_detail[row_num]:
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left", vertical="center")

    _auto_width(ws_detail)

    wb.save(output_path)
    logger.info("Excel report saved: %s (%d rows)", output_path, len(violations))
